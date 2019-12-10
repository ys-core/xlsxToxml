

import os
from os import path
import xlrd     # module to deal excel
from xml.dom.minidom import Document
from openpyxl import load_workbook

from tkinter import *
from tkinter import messagebox
root = Tk()
root.withdraw()


xmlNumbers = 0
rootdir = './'   # crrent directory
ecos_data_dir = 'ECOS_Data'
Feature_Code = 'Feature_Code'


original_path = path.dirname(__file__)
# print(original_path)


os.chdir(ecos_data_dir)     #change directory ECOS_Data
# print(os.getcwd())

list = os.listdir(os.getcwd())
# print(list)


for line in list:
    # print(line)
    filepath = os.path.join(rootdir, line)
    # print(filepath)
    #if os.path.isdir(filepath):
        # print ("dir:" + filepath)
    if os.path.isfile(filepath):
        # print ("file:" + filepath)
        if '.xlsx' in filepath:
            # print(str(filepath)[2:])

            wb = load_workbook(filepath)
            sheet1Name = wb.sheetnames[0]
            ws = wb[sheet1Name]
            date_value = str(ws['B6'].value)
            date_value = date_value.replace(' ','T',1)
            date_ = date_value + '+08:00'




            book = xlrd.open_workbook(filepath)
            # sheets number in this workbook
            count = len(book.sheets())
            if count == 6:    #deal with workbook which contains just 6 worksheets

                doc = Document()  # create the Doc element
                ROOT = doc.createElement('AsBuild')  # create the root element
                doc.appendChild(ROOT)
                """
                ///////////////////////////////////////    Version
                """
                Version = doc.createElement('Version')
                ROOT.appendChild(Version)
                version_value = doc.createTextNode('1.0')
                Version.appendChild(version_value)


                sheet5 = book.sheet_by_index(4)  #read the 5th sheet
                #print(str(sheet5.col(1)[2]))

                #print (str(sheet5.col(1)[2])[6:].lstrip()[0:7])      # B3 cell

                """
                /////////////////////////////////////    GeneralInfo
                """
                GeneralInfo = doc.createElement('GeneralInfo')
                ROOT.appendChild(GeneralInfo)

                ProdID = doc.createElement("ProID")
                GeneralInfo.appendChild(ProdID)
                value = doc.createTextNode(str(sheet5.col(1)[2])[6:].lstrip()[0:7])
                ProdID.appendChild(value)

                Plant = doc.createElement("Plant")
                GeneralInfo.appendChild(Plant)
                value = doc.createTextNode('CS')
                Plant.appendChild(value)

                ProdYear = doc.createElement("ProdYear")
                GeneralInfo.appendChild(ProdYear)

                start = str(sheet5.col(1)[2]).find('---') + 3
                end = len(str(sheet5.col(1)[2]))-2
                #print(str(sheet5.col(1)[2]))
                #print(str(sheet5.col(1)[2])[start:end].rstrip()[:-1])
                VehicleIdentNumber = doc.createElement("VehicleIdentNumber")
                GeneralInfo.appendChild(VehicleIdentNumber)
                value = doc.createTextNode(str(sheet5.col(1)[2])[start:end].rstrip()[:-1])
                VehicleIdentNumber.appendChild(value)

                end = str(sheet5.col(1)[2])[6:].lstrip().find('---')
                # print(str(sheet5.col(1)[2])[6:].lstrip()[8:end])
                Mode = doc.createElement("Mode")
                GeneralInfo.appendChild(Mode)
                value = doc.createTextNode(str(sheet5.col(1)[2])[6:].lstrip()[8:end])
                Mode.appendChild(value)

                # return to root directory, then change directory to Feature_Code
                current_directory = os.getcwd()
                # print(current_directory)
                os.chdir(path.pardir)
                os.chdir(Feature_Code)
                newFilePath = os.getcwd()+'\\'+filepath[2:]
                # print(newFilePath)
                # print(os.path.pardir)
                # wb_ = load_workbook(newFilePath)
                # sheet1Name_ = wb_.sheetnames[0]
                # ws_ = wb_[sheet1Name_]
                # value_ = str(ws_['A2'].value)

                book_ = xlrd.open_workbook(filepath)
                sheet1_ = book_.sheet_by_index(0)
                value_ = str(sheet1_.col(0)[1])[6:-1]
                # print(value_)

                # print(value_)
                FeatureCodes = doc.createElement("FeatureCodes")
                GeneralInfo.appendChild(FeatureCodes)
                value = doc.createTextNode(value_)
                FeatureCodes.appendChild(value)

                os.chdir(path.pardir)
                os.chdir(ecos_data_dir)




                sheet1 = book.sheet_by_index(0)  # read the 1st sheet



                """
                ////////////////////////////////////     TesterInfo
                """
                TesterInfo = doc.createElement("TesterInfo")
                ROOT.appendChild(TesterInfo)

                TesterId = doc.createElement("TesterId")
                TesterInfo.appendChild(TesterId)
                start = 6
                end = len(str(sheet1.col(1)[3])) - 1
                value = doc.createTextNode(str(sheet1.col(1)[3])[start:end])
                TesterId.appendChild(value)

                Date = doc.createElement("Date")            # date type , ctype == 3
                TesterInfo.appendChild(Date)
                value = doc.createTextNode(date_)
                #print(sheet1.col(1)[5])
                Date.appendChild(value)




                sheet4 = book.sheet_by_index(3)  #  read the 4th sheet
                #print(len(sheet4.col_values(3)))  # 4th column data in 4th sheet
                vpdataList = []
                for index in range(len(sheet4.col_values(3))):
                    #print(sheet4.col(3)[index])
                    if('VPDATA' in str(sheet4.col(3)[index])):
                        vpdataList.append(str(sheet4.col(3)[index])[14:])
                vpdataList.sort()
                #print(len(vpdataList))
                p = vpdataList[0].find('"')


                temp = vpdataList[0][0:p-2]
                for i, item in enumerate(vpdataList):
                    #print(i,item)
                    start = str(vpdataList[i]).find(";") + 2
                    end = str(vpdataList[i]).find("}") - 2
                    token_node_value = str(vpdataList[i])[start:end]
                    #print(token_node_value)
                    p = vpdataList[i].find('"')

                    if i == 0:
                        """
                        ////////////////////////////////////     Feature, Tokes 
                        """
                        Feature = doc.createElement("Feature")
                        ROOT.appendChild(Feature)
                        Feature.setAttribute("name", temp)
                        Tokens = doc.createElement("Tokens")
                        Feature.appendChild(Tokens)

                    current = vpdataList[i][0:p-2]
                    if current == temp:
                        Token = doc.createElement("Token")
                        Tokens.appendChild(Token)
                        token_name = vpdataList[i][p-2:p]
                        Token.setAttribute("name",token_name)
                        value = doc.createTextNode(token_node_value)
                        Token.appendChild(value)
                    else:

                        Feature = doc.createElement("Feature")
                        ROOT.appendChild(Feature)
                        Feature.setAttribute("name", current)
                        Tokens = doc.createElement("Tokens")
                        Feature.appendChild(Tokens)

                        Token = doc.createElement("Token")
                        Tokens.appendChild(Token)
                        token_name = vpdataList[i][p - 2:p]
                        Token.setAttribute("name", token_name)
                        value = doc.createTextNode(token_node_value)
                        Token.appendChild(value)

                        temp = current




                # print(cwd)
                # print(original_path)
                os.chdir(path.pardir)
                cwd = os.getcwd()  # create new xml directory to save all generated .xml fiels
                # print(os.getcwd())
                #
                new_cwd = '\\xml'
                # print(cwd + new_cwd)
                if not os.path.exists(cwd + new_cwd):
                    os.mkdir(cwd + new_cwd)
                    os.chdir(cwd + new_cwd)
                else:
                    os.chdir(cwd + new_cwd)
                # print(os.getcwd())

                xlxsFileName = str(filepath)[2:]
                end = xlxsFileName.find(".")
                xmlFileName = xlxsFileName[0:end] + '.xml'
                # print(xmlFileName)
                f = open(xmlFileName, 'w',encoding="UTF-8")
                # f.write(doc.toprettyxml(indent = '\t', newl = '\n', encoding = 'utf-8'))
                #doc.writexml(f, indent='\t', newl='\n', addindent='\t', encoding='utf-8')
                doc.writexml(f, indent='', newl='\n', addindent='\t', encoding='UTF-8')
                f.close()
                xmlNumbers = xmlNumbers + 1
                os.chdir(path.pardir)
                os.chdir(ecos_data_dir)   # return back to parent directory for next loop
                # print(os.getcwd())

resultMessage = str(xmlNumbers) +" files have been transfered !!"    # tell user the result
txt = messagebox.showinfo("Result",resultMessage)


if txt=="ok":
    root.destroy()
    root.mainloop()